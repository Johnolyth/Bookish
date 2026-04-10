# -------------------------------------------
# -------------------------------------------
# ----------------BOOKISH APP----------------
# -------------------------------------------
# -------------------------------------------
# -------------------------------------------
# -----------------MISC Setup----------------


import json
import os
from datetime import datetime
from openpyxl import load_workbook
workbook_path = "bookish_test.xlsx"
workbook = load_workbook(workbook_path)
books_sheet = workbook["Books"]
logs_sheet = workbook["Logs"]
# -----------------FUNCTIONS-----------------


def save_workbook():
    workbook.save(workbook_path)
    print("✅ Changes saved to workbook.\n")


def log_note(title):

    print(f"Logging a note for: {title}")

    page = input("Enter page number (optional): ").strip()
    if not page:
        page = "--"

    note = input(f"Enter your note for '{title}': ").strip()
    if not note:
        print("No note entered.\n")
        return

    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    logs_sheet.append([title, page, note, timestamp])
    print(f"📝 Note logged for '{title}' at {timestamp}.\n")


def mark_owned(title):
    for row in books_sheet.iter_rows(min_row=2):
        cell_title = row[0].value
        if cell_title and cell_title.strip().lower() == title.strip().lower():
            row[4].value = "yes"
            print(f"📚 '{title}' marked as owned.\n")
            break
    else:
        print(f"⚠️ Title '{title}' not found in Books sheet.\n")
        workbook.save(workbook_path)


def mark_read(title):
    for row in books_sheet.iter_rows(min_row=2):
        cell_title = row[0].value
        if cell_title and cell_title.strip().lower() == title.strip().lower():
            row[5].value = "yes"
            print(f"📖 '{title}' marked as read.\n")
            break

    else:
        print(f"⚠️ Title '{title}' not found in Books sheet.\n")


def search_by_title():
    search_term = input("Enter a book title (or part of it): ").strip().lower()

    print("\nResults:")
    found = False
    for row in books_sheet.iter_rows(min_row=2, values_only=True):
        title = str(row[0]).lower()
        if search_term in title:
            found = True
            print(f"- {row[0]} by {row[1]}")

            choice = input(
                "\nWould you like to (L)og a note, mark as (O)wned, mark as (R)ead, or (S)kip? ").strip().lower()

            if choice == "l":
                log_note(row[0])
                save_workbook()
            elif choice == "o":
                mark_owned(row[0])
                save_workbook()
            elif choice == "r":
                mark_read(row[0])
                save_workbook()
            elif choice == "s":
                print("Skipped.\n")
            else:
                print("Invalid choice.\n")

    if not found:
        print("No matches found")


def search_by_author():
    search_term = input(
        "Enter an author name (or part of it): ").strip().lower()

    matching_books = []

    for row in books_sheet.iter_rows(min_row=2, values_only=True):
        author = str(row[1]).lower()
        if search_term in author:
            matching_books.append((row[0], row[1]))

    if not matching_books:
        print("No books found by that author.\n")
        return

    print("\nBooks by matching author(s):")
    for idx, (title, author) in enumerate(matching_books, start=1):
        print(f"{idx}. {title} by {author}")

    try:
        selection = input(
            "Select a book by number or press Enter to cance: ").strip()
        if selection == "":
            print("Selection cancelled.\n")
            return

        selection = int(selection)
        if 1 <= selection <= len(matching_books):
            selected_title = matching_books[selection - 1][0]
            print(f"You selected: {selected_title}")

            choice = input(
                "Would you like to (L)og a note, mark as (O)wned, mark as (R)ead, or (S)kip? ").strip().lower()

            if choice == "l":
                log_note(selected_title)
                save_workbook()

            elif choice == "o":
                mark_owned(selected_title)
                save_workbook()

            elif choice == "r":
                mark_read(selected_title)
                save_workbook()

            elif choice == "s":
                print("Skipped.\n")

            else:
                print("Invalid choice.\n")

        else:
            print("Invalid selection.\n")

    except ValueError:
        print("Invalid input. Please enter a number.\n)")


def XXsearch_by_authorXX():
    search_term = input(
        "Enter an author name (or part of it): ").strip().lower()
    print("\nBooks by matching authos:")

    found = False
    for row in books_sheet.iter_rows(min_row=2, values_only=True):
        author = str(row[1]).lower()
        title = row[0]
        if search_term in author:
            found = True
            print(f"- {title} by {row[1]}")
    if not found:
        print("No matches found.")


def main_menu():
    print("--- Main Menu ---")
    print("\n-Search\n-Quit")


def search_menu():
    print("--- Search Menu ---")
    print("\nAuthor (search by author)\nTitle (search by title)\nBack")

# -------------------------------------------


main_command = ""
filename = "book_notes.json"

if os.path.exists(filename):
    with open(filename, "r") as file:
        book_notes = json.load(file)

else:
    book_notes = {}


# -----------------MAIN BLOCK----------------

while main_command.strip().lower() != "quit":
    main_menu()
    main_command = input("> ")

    while main_command.strip().lower() != "back":
        search_menu()
        main_command = input("> ")

        if main_command.strip().lower() == "title":
            search_by_title()

        elif main_command.strip().lower() == "author":
            search_by_author()
